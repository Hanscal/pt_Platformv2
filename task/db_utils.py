#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
@Time    : 2020/5/27 11:41 上午
@Author  : hcai
@Email   : hua.cai@unidt.com
"""

import os
import sys
import xlrd
import openpyxl
from openpyxl.utils import get_column_letter
from .models import DataItem

file_dir = os.path.dirname(__file__)
UPLOAD_ROOT = os.path.join(file_dir,'upload')
if not os.path.isdir(UPLOAD_ROOT):
    os.mkdir(UPLOAD_ROOT)

def process_data(data_list):
    res = []
    for data in data_list:
        data = data.decode('utf-8') if isinstance(data,bytes) else data
        line = data.strip().split()
        print(line)
        res.append(line[:])
    return res

def load_file(file_dir, extension):
    """extention in jpg or txt supported"""
    file_list = os.listdir(file_dir)
    file_path_list = []
    for file in file_list:
        if file.split('.')[-1] not in extension:
            continue
        file_path = os.path.join(file_dir, file)
        file_path_list.append(file_path)
    return file_path_list

def save_file(filename, data_list):
    file_path = os.path.join(UPLOAD_ROOT, filename)
    with open(file_path, 'w') as fw:
        for data in data_list:
            data = data.decode('utf-8') if isinstance(data,bytes) else data
            line = data.strip().split()
            fw.write('\t'.join(line) + '\n')

def process_excel_file(file):
    res = []
    wb = openpyxl.load_workbook(file)  # 打开文件
    sheetnames_list = wb.sheetnames  # 获取所有表格名字
    sheet1 = wb.get_sheet_by_name(sheetnames_list[0])  # 通过索引获取表格
    # print(sheet1.name, sheet1.nrows, sheet1.ncols)

    for row in sheet1.iter_rows():
        row_data = list()
        for cell in row:
            row_data.append(str(cell.value))
        res.append(row_data)
    return res

def process_txt_file(file_path):
    """
    mid	 img_name	txt	 task
    1	  aaa.jpg	11	 1
    2	  bbb.jpg	12	 1
    3	  ccc.jpg	13	 1
    :param file_path:
    :return:
    """
    data_list = []
    with open(file_path, 'r') as f:
        content = f.readlines()
        for line in content:
            line_list = line.strip().split()
            data_list.append(line_list[:])
    return data_list

def write_data(dataitem, data_list):
    for data in data_list:
        print("save image name: {}".format(data[1]))
        data_item = dataitem()
        data_item.mid = data[0]
        data_item.img_name = data[1]
        data_item.txt = data[2]
        data_item.task = data[3]
        data_item.save()